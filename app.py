import streamlit as st
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os
import tempfile
from io import BytesIO
from PIL import Image as PILImage

# Custom CSS for modern dark mode design
st.markdown("""
<style>
    /* General modern styling with dark mode */
    .stApp {
        background-color: #121212;
        color: #EDEDED;
    }
    h1, h2, h3, .stMarkdown, .stText {
        color: #EDEDED;
    }
    /* Sidebar styling */
    [data-testid="stSidebar"] {
        background-color: #1F1F1F;
        border-right: 1px solid #333333;
    }
    [data-testid="stSidebar"] .stButton > button {
        background-color: #2196F3;
        color: #FFFFFF;
        border-radius: 8px;
    }
    /* Expander styling */
    .stExpander {
        border: 1px solid #333333;
        border-radius: 8px;
        background-color: #1F1F1F;
    }
    .stExpander summary {
        color: #EDEDED;
        font-weight: bold;
    }
    /* Button styling */
    .stButton > button {
        background-color: #4CAF50;
        color: #FFFFFF;
        border-radius: 8px;
        padding: 10px 20px;
        font-size: 16px;
        transition: background-color 0.3s;
    }
    .stButton > button:hover {
        background-color: #388E3C;
    }
    /* Download button */
    .stDownloadButton > button {
        background-color: #2196F3;
        color: #FFFFFF;
        border-radius: 8px;
        padding: 10px 20px;
        font-size: 16px;
        transition: background-color 0.3s;
    }
    .stDownloadButton > button:hover {
        background-color: #1976D2;
    }
    /* Dataframe styling */
    .stDataFrame {
        border: 1px solid #333333;
        border-radius: 8px;
        background-color: #1F1F1F;
    }
    .stDataFrame th, .stDataFrame td {
        color: #EDEDED;
    }
    /* Slider styling */
    .stSlider .stMarkdown {
        color: #EDEDED;
    }
    /* Progress bar */
    .stProgress > div > div > div > div {
        background-color: #4CAF50;
    }
    /* Info, success, warning */
    .stAlert {
        border-radius: 8px;
    }
    .stInfo {
        background-color: #1F1F1F;
        color: #BBDEFB;
    }
    .stSuccess {
        background-color: #1F1F1F;
        color: #A5D6A7;
    }
    .stWarning {
        background-color: #1F1F1F;
        color: #FFCC80;
    }
</style>
""", unsafe_allow_html=True)

# Function to download and compress image
def download_image(url, temp_dir):
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        # Load image with Pillow
        img_data = BytesIO(response.content)
        img = PILImage.open(img_data)
        # Convert to RGB if necessary (e.g., for PNG with transparency)
        if img.mode in ('RGBA', 'LA'):
            img = img.convert('RGB')
        # Resize to max 800x800 pixels
        img.thumbnail((800, 800), PILImage.Resampling.LANCZOS)
        # Save compressed image as JPEG
        file_name = os.path.basename(url.split('?')[0])
        if not file_name.lower().endswith(('.jpg', '.jpeg')):
            file_name = file_name + '.jpg'
        file_path = os.path.join(temp_dir, file_name)
        img.save(file_path, 'JPEG', quality=85, optimize=True)
        return file_path
    except Exception as e:
        st.warning(f"Error downloading or processing {url}: {e}")
        return None

# Function to convert CSV to XLSX
def csv_to_xlsx(file_bytes, temp_dir):
    df = pd.read_csv(BytesIO(file_bytes))
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col_idx, column in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx).value = column
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx).value = value
    temp_xlsx = os.path.join(temp_dir, "converted.xlsx")
    wb.save(temp_xlsx)
    return temp_xlsx, df

# Sidebar for navigation and settings
with st.sidebar:
    st.title("ImageNest Settings")
    st.markdown("Customize row and column sizes")
    
    # User controls for sizes
    row_height = st.slider("Row Height (points)", min_value=40, max_value=400, value=300, step=10)
    col_width = st.slider("Column Width (characters)", min_value=44, max_value=100, value=44, step=1)
    
    st.markdown("---")
    st.caption("ImageNest v1.0")
    st.caption("Made with ❤️ by Hassanelb")

# Main content with collapsible expanders
st.title("ImageNest")
st.title("CSV/Excel Image Embedder")
st.markdown("Embed images from URLs into your spreadsheets for Canva compatibility. Upload a file, select a column, and download the result.")

# Expander for File Upload
with st.expander("Step 1: Upload File", expanded=True):
    uploaded_file = st.file_uploader("Choose a CSV or Excel file", type=["csv", "xlsx"], help="Upload your file here.")

# Process if file is uploaded
if uploaded_file is not None:
    file_bytes = uploaded_file.getvalue()
    
    # Determine file type and load sheet names
    if uploaded_file.name.endswith('.csv'):
        sheet_names = ["Sheet1"]
        df = pd.read_csv(BytesIO(file_bytes))
    else:
        xl = pd.ExcelFile(BytesIO(file_bytes))
        sheet_names = xl.sheet_names
        df = None
    
    # Expander for Sheet Selection
    with st.expander("Step 2: Select Sheet", expanded=True):
        sheet_name = st.selectbox("Select the sheet", sheet_names, help="Choose the sheet to process.")
    
    if sheet_name:
        # Load preview data
        if uploaded_file.name.endswith('.csv'):
            preview_df = df
        else:
            preview_df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
        
        # Expander for File Preview
        with st.expander("Step 3: File Preview", expanded=True):
            st.dataframe(preview_df.head(10), use_container_width=True)
        
        # Expander for Column Selection
        with st.expander("Step 4: Select Column", expanded=True):
            column_headers = preview_df.columns.tolist()
            image_col_name = st.selectbox("Select the column with image URLs", column_headers, help="Choose the column containing image URLs.")
        
        # Expander for Insert Images
        with st.expander("Step 5: Insert Images", expanded=True):
            if st.button("Start Processing", use_container_width=True):
                with st.spinner("Processing images... Please wait."):
                    with tempfile.TemporaryDirectory() as temp_dir:
                        # Handle CSV or XLSX
                        if uploaded_file.name.endswith('.csv'):
                            temp_xlsx, _ = csv_to_xlsx(file_bytes, temp_dir)
                        else:
                            temp_xlsx = os.path.join(temp_dir, "uploaded.xlsx")
                            with open(temp_xlsx, "wb") as f:
                                f.write(file_bytes)
                        
                        # Load workbook
                        wb = load_workbook(temp_xlsx)
                        ws = wb[sheet_name]
                        
                        # Insert new column
                        image_col = column_headers.index(image_col_name) + 1
                        new_image_col = image_col + 1
                        ws.insert_cols(new_image_col)
                        ws.cell(row=1, column=new_image_col).value = "Embedded Image"
                        
                        # Process images with progress bar
                        num_rows = ws.max_row - 1
                        progress_bar = st.progress(0.0)
                        processed = 0
                        
                        for row in range(2, ws.max_row + 1):
                            url_cell = ws.cell(row=row, column=image_col)
                            url = url_cell.value
                            if url:
                                image_path = download_image(url, temp_dir)
                                if image_path:
                                    img = Image(image_path)
                                    img.width = 300  # Fixed size for Canva compatibility
                                    img.height = 300
                                    # Center image in cell
                                    ws.add_image(img, ws.cell(row=row, column=new_image_col).coordinate)
                                    ws.row_dimensions[row].height = row_height
                                else:
                                    ws.cell(row=row, column=new_image_col).value = "Failed to download"
                            
                            processed += 1
                            progress_bar.progress(processed / num_rows)
                        
                        # Adjust column width
                        ws.column_dimensions[ws.cell(row=1, column=new_image_col).column_letter].width = col_width
                        
                        # Save to BytesIO and normalize
                        temp_output = BytesIO()
                        wb.save(temp_output)
                        temp_output.seek(0)
                        wb_temp = load_workbook(temp_output)
                        output = BytesIO()
                        wb_temp.save(output)
                        output.seek(0)
                
                # Success message and download
                st.success("Processing complete! Download the modified file.")
                st.download_button(
                    label="Download Modified Excel",
                    data=output,
                    file_name="data_with_images.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
else:
    st.info("Upload a file to begin.")
